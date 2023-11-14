import openpyxl
import os
import pandas as pd

a = []
p = []
file = "E:\\ABC"
files = os.listdir('E:\ABC')
for i in files:
    if i.endswith('.xlsx'):
        print(f"{i=}")
        y = openpyxl.load_workbook(os.path.join(file, i))
        for sheet in y:
            b = sheet.title
            if b not in a:
                a.append(b)
                # with pd.ExcelWriter(b + .xlsx) as Writer:

                print(f"{b=}")
                print(f"{a=}")
outfile = 'E:/ABC/New folder'
outfile2 = "E:\\ABC\\New folder"
os.remove(os.path.join(outfile, "Output.xlsx"))
os.rmdir(outfile)
os.mkdir(outfile, mode=0o777)
ws = openpyxl.Workbook()
ws = ws.save(os.path.join(outfile, "Output.xlsx"))
wb = openpyxl.load_workbook(os.path.join(outfile, "Output.xlsx"))
# aa=openpyxl.load_workbook(os.path.join(outfile2, "Output.xlsx"))
# for j in a:
#    for i in files:
#        if i.endswith('.xlsx'):
for j in a:
    for i in files:
        if i.endswith('.xlsx'):
            xl = pd.ExcelFile(os.path.join(file, i))
            if j in openpyxl.load_workbook(os.path.join(file, i)).sheetnames:
                if j not in p:
                    p.append(j)
                    print(f"{p=}")
                    df = xl.parse(j)
                    # last_row = openpyxl.load_workbook(os.path.join(file, i))[j].max_row+1
                    # last_row = wb[j].max_row+1
                    sheet_exists = wb.sheetnames
                    print(f"{sheet_exists=}")
                    print(f"{j=}")
                    # try:
                    # if j not in p:
                    with pd.ExcelWriter(os.path.join(outfile, "Output.xlsx"), engine='openpyxl', mode='a', if_sheet_exists=None) as writer:
                        df.to_excel(writer, startrow=0, sheet_name=j, index=False)
                        # wb.SaveAs(os.path.join(outfile, "Output.xlsx"))
                        # writer.Save()
                        # wb.Close(True)
                elif j in p:
                     print("{j2=}"+ str(j))
                     print(f"{j=}")
                # last_row = wb.get_sheet_by_name(j).max_row + 1
                     last_row = openpyxl.load_workbook("E:\\ABC\\New folder\\Output.xlsx")[j].max_row
                     print(last_row)
                     with pd.ExcelWriter(os.path.join(outfile, "Output.xlsx"), engine='openpyxl', mode='a',
                                    if_sheet_exists='overlay') as writer:
                         df.to_excel(writer, sheet_name=j, startrow=last_row, index=False, header=False)
                        #read existing sheet's all columns, and iterate sheet columns and if matches append or new column
                    # wb.SaveAs(os.path.join(outfile, "Output.xlsx"))
                    # wb.Save()
                    # writer.Save()
                    # wb.Close(True)
                    # writer.close()

            # except PermissionError:
            # print(f'Error: Permission denied while trying to access {aa}.')
            # except Exception as e:
            # print(f'An error occurred: {e}')
# writer.close()